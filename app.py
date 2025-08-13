from flask import Flask, request, render_template, send_file, flash, redirect, url_for
import PyPDF2
import pdfplumber
import os
import re
from openpyxl import load_workbook
import tempfile
import io
from datetime import datetime
from excel_to_image import excel_to_image_with_cropping

app = Flask(__name__)
app.secret_key = 'your-secret-key-here'
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max file size

# Ensure upload directory exists
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

def get_pdf_crop_boundaries(pdf_path):
    """Determine crop boundaries for the first column based on file type"""
    filename = os.path.basename(pdf_path).lower()
    
    # Based on analysis, set crop boundaries for different PDF types
    if 'arr' in filename:
        return {'x0': 0, 'x1': 229}  # ARR files
    elif 'dep' in filename:
        return {'x0': 0, 'x1': 197}  # DEP files  
    elif 'gih' in filename:
        return {'x0': 0, 'x1': 220}  # GIH files
    else:
        # Default - try to detect automatically
        return {'x0': 0, 'x1': 220}  # Conservative default

def extract_room_numbers_from_pdf(pdf_path):
    """Extract room numbers from PDF file using first column cropping"""
    room_numbers = set()
    
    # Get crop boundaries for this PDF type
    crop = get_pdf_crop_boundaries(pdf_path)
    print(f"Using crop boundaries for {os.path.basename(pdf_path)}: x={crop['x0']}-{crop['x1']}")
    
    try:
        with pdfplumber.open(pdf_path) as pdf:
            for page_num, page in enumerate(pdf.pages, 1):
                # Crop page to only the first column
                cropped_page = page.crop((crop['x0'], 0, crop['x1'], page.height))
                
                # Extract text from cropped area only
                text = cropped_page.extract_text()
                if text:
                    print(f"Page {page_num} cropped text preview: {text[:100]}...")
                    
                    # Look for room numbers in format like 0211, 0214, 1011, etc.
                    # Use 3-4 digit pattern to catch both formats
                    room_pattern = r'\b(\d{3,4})\b'
                    matches = re.findall(room_pattern, text)
                    
                    print(f"Found potential room numbers: {matches[:10]}...")
                    
                    # Filter to keep only valid room numbers
                    for match in matches:
                        if len(match) >= 3 and match.isdigit():
                            # Convert to format like 211, 214 (remove leading zero if exists)
                            room_num = match.lstrip('0') if match.startswith('0') else match
                            if room_num and len(room_num) >= 3:  # Valid room numbers should be at least 3 digits
                                room_int = int(room_num)
                                # More lenient filtering since we're only looking at first column
                                if 100 <= room_int <= 9999:
                                    room_numbers.add(room_int)
                                    
    except Exception as e:
        print(f"Error extracting from PDF {pdf_path}: {str(e)}")
        # Fallback to old method if cropping fails
        try:
            print("Falling back to full-text extraction...")
            with pdfplumber.open(pdf_path) as pdf:
                for page in pdf.pages:
                    text = page.extract_text()
                    if text:
                        room_pattern = r'\b(\d{3,4})\b'
                        matches = re.findall(room_pattern, text)
                        
                        for match in matches:
                            if len(match) >= 3 and match.isdigit():
                                room_num = match.lstrip('0') if match.startswith('0') else match
                                if room_num and len(room_num) >= 3:
                                    room_int = int(room_num)
                                    if 100 <= room_int <= 9999 and not (2500 <= room_int <= 2600):
                                        room_numbers.add(room_int)
        except Exception as e2:
            print(f"Fallback extraction also failed: {str(e2)}")
    
    print(f"Final extracted room numbers ({len(room_numbers)}): {sorted(list(room_numbers)[:10])}...")
    return room_numbers

def update_excel_template(template_path, arr_rooms, dep_rooms, gih_rooms, output_path):
    """Update Excel template with room data"""
    try:
        wb = load_workbook(template_path)
        ws = wb.active
        
        print(f"Processing rooms:")
        print(f"ARR: {sorted(arr_rooms)}")
        print(f"DEP: {sorted(dep_rooms)}")
        print(f"GIH: {sorted(gih_rooms)}")
        
        # Based on analysis, headers are in row 4, room data starts from row 5
        header_row = 4
        
        # Find header columns for each section
        header_positions = []
        for col in range(1, ws.max_column + 1):
            header_cell = ws.cell(row=header_row, column=col)
            if header_cell.value:
                header_val = str(header_cell.value).strip()
                if header_val == 'Room':
                    # Found a room column, next columns should be OD, DO, ARR, NOTE
                    room_col = col
                    od_col = col + 1 if col + 1 <= ws.max_column else None
                    do_col = col + 2 if col + 2 <= ws.max_column else None
                    arr_col = col + 3 if col + 3 <= ws.max_column else None
                    
                    # Verify column headers
                    od_header = ws.cell(row=header_row, column=od_col).value if od_col else None
                    do_header = ws.cell(row=header_row, column=do_col).value if do_col else None
                    arr_header = ws.cell(row=header_row, column=arr_col).value if arr_col else None
                    
                    if (od_header and 'OD' in str(od_header) and 
                        do_header and 'DO' in str(do_header) and 
                        arr_header and 'ARR' in str(arr_header)):
                        
                        header_positions.append({
                            'room_col': room_col,
                            'od_col': od_col,
                            'do_col': do_col,
                            'arr_col': arr_col
                        })
        
        print(f"Found {len(header_positions)} room sections")
        
        # Process each row starting from row 5
        for row_num in range(5, ws.max_row + 1):
            for section in header_positions:
                room_cell = ws.cell(row=row_num, column=section['room_col'])
                if room_cell.value:
                    try:
                        room_value = str(room_cell.value).strip()
                        if room_value.isdigit():
                            # Convert room number (handle formats like 0211 -> 211)
                            room_num = int(room_value.lstrip('0')) if room_value.startswith('0') else int(room_value)
                            
                            # Mark appropriate columns
                            if room_num in gih_rooms and section['od_col']:
                                ws.cell(row=row_num, column=section['od_col'], value='x')
                                print(f"Marked room {room_num} in OD column")
                            
                            if room_num in dep_rooms and section['do_col']:
                                ws.cell(row=row_num, column=section['do_col'], value='x')
                                print(f"Marked room {room_num} in DO column")
                            
                            if room_num in arr_rooms and section['arr_col']:
                                ws.cell(row=row_num, column=section['arr_col'], value='x')
                                print(f"Marked room {room_num} in ARR column")
                    except ValueError:
                        continue  # Skip non-numeric room values
        
        wb.save(output_path)
        print(f"Excel file saved to: {output_path}")
        return True
    except Exception as e:
        print(f"Error updating Excel template: {str(e)}")
        import traceback
        traceback.print_exc()
        return False


@app.route('/')
def index():
    return render_template('index.html')

@app.route('/process', methods=['POST'])
def process_files():
    try:
        # Check if all required files are uploaded
        if 'arr_file' not in request.files or 'dep_file' not in request.files or 'gih_file' not in request.files:
            flash('Vui lòng upload đầy đủ 3 file: ARR, DEP, và GIH')
            return redirect(url_for('index'))
        
        arr_file = request.files['arr_file']
        dep_file = request.files['dep_file']
        gih_file = request.files['gih_file']
        
        if arr_file.filename == '' or dep_file.filename == '' or gih_file.filename == '':
            flash('Vui lòng chọn đầy đủ 3 file')
            return redirect(url_for('index'))
        
        # Check output format
        output_format = request.form.get('output_format', 'excel')
        
        # Save uploaded files
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        arr_path = os.path.join(app.config['UPLOAD_FOLDER'], f'arr_{timestamp}.pdf')
        dep_path = os.path.join(app.config['UPLOAD_FOLDER'], f'dep_{timestamp}.pdf')
        gih_path = os.path.join(app.config['UPLOAD_FOLDER'], f'gih_{timestamp}.pdf')
        
        arr_file.save(arr_path)
        dep_file.save(dep_path)
        gih_file.save(gih_path)
        
        # Extract room numbers from each PDF
        arr_rooms = extract_room_numbers_from_pdf(arr_path)
        dep_rooms = extract_room_numbers_from_pdf(dep_path)
        gih_rooms = extract_room_numbers_from_pdf(gih_path)
        
        print(f"ARR rooms: {arr_rooms}")
        print(f"DEP rooms: {dep_rooms}")
        print(f"GIH rooms: {gih_rooms}")
        
        # Update Excel template
        template_path = 'template.xlsx'
        output_excel_path = os.path.join(app.config['UPLOAD_FOLDER'], f'result_{timestamp}.xlsx')
        
        success = update_excel_template(template_path, arr_rooms, dep_rooms, gih_rooms, output_excel_path)
        
        if not success:
            flash('Có lỗi xảy ra khi xử lý file Excel')
            return redirect(url_for('index'))
        
        # Convert to image if requested
        if output_format == 'image':
            output_image_path = os.path.join(app.config['UPLOAD_FOLDER'], f'result_{timestamp}.png')
            image_success = excel_to_image_with_cropping(output_excel_path, output_image_path)
            
            if image_success:
                return send_file(output_image_path, 
                                as_attachment=True, 
                                download_name=f'housekeeping_report_{timestamp}.png',
                                mimetype='image/png')
            else:
                flash('Có lỗi khi tạo file ảnh, tải file Excel thay thế')
        
        # Send the Excel file (default or fallback)
        return send_file(output_excel_path, 
                        as_attachment=True, 
                        download_name=f'housekeeping_report_{timestamp}.xlsx',
                        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        
    except Exception as e:
        flash(f'Có lỗi xảy ra: {str(e)}')
        return redirect(url_for('index'))

if __name__ == '__main__':
    app.run(debug=True)
