#!/usr/bin/env python3
"""
Flask Web Server cho Hotel Room Classification System
Cho phép upload PDF files và xử lý tự động
"""

from flask import Flask, render_template, request, jsonify, redirect, url_for, send_file
import os
import subprocess
import tempfile
import re
import pandas as pd
from datetime import datetime, timedelta
from werkzeug.utils import secure_filename
from pdf2image import convert_from_path
from PIL import Image
import pytesseract
import zipfile
import shutil

# Configure tesseract path for different environments
# Try to find tesseract executable in common locations
tesseract_paths = [
    '/usr/bin/tesseract',        # Linux/Docker
    '/opt/homebrew/bin/tesseract', # macOS with Homebrew (Apple Silicon)
    '/usr/local/bin/tesseract',   # macOS with Homebrew (Intel)
    'tesseract'                   # System PATH
]

for path in tesseract_paths:
    if path == 'tesseract' or shutil.which(path):
        try:
            pytesseract.pytesseract.tesseract_cmd = path
            # Test if it works
            result = subprocess.run([path, '--version'], capture_output=True, text=True)
            if result.returncode == 0:
                print(f"✅ Tesseract found at: {path}")
                print(f"   Version: {result.stdout.split()[1] if result.stdout else 'Unknown'}")
                break
        except Exception as e:
            print(f"❌ Failed to use tesseract at {path}: {e}")
            continue
else:
    print("⚠️ Warning: Tesseract not found in any common locations")

app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max file size
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', 'dev-secret-key-change-in-production')

# Disable CSRF for file uploads if needed
app.config['WTF_CSRF_ENABLED'] = False

# Production-friendly upload folder
UPLOAD_FOLDER = os.path.join(os.getcwd(), 'temp_uploads')
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

ALLOWED_EXTENSIONS = {'pdf', 'PDF'}
IMAGE_EXTENSIONS = {'jpg', 'jpeg', 'png', 'JPG', 'JPEG', 'PNG'}

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1] in ALLOWED_EXTENSIONS

def allowed_image_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1] in IMAGE_EXTENSIONS

def allowed_zip_file(filename):
    return filename.lower().endswith('.zip')

def extract_text_from_image(image_path):
    """Extract text from image using OCR (pytesseract)"""
    try:
        print(f"🔍 Processing image: {os.path.basename(image_path)}")
        
        # Open image with PIL
        image = Image.open(image_path)
        print(f"   Image size: {image.size}, mode: {image.mode}")
        
        # Convert to RGB if necessary
        if image.mode != 'RGB':
            image = image.convert('RGB')
            print(f"   Converted to RGB mode")
        
        # Use pytesseract to extract text with better configuration for hotel data
        # config options: preserve_interword_spaces to maintain table structure
        custom_config = r'--oem 3 --psm 6 -c preserve_interword_spaces=1'
        text = pytesseract.image_to_string(image, lang='eng', config=custom_config)
        
        lines_found = len([line for line in text.split('\n') if line.strip()])
        print(f"   ✅ OCR completed: {lines_found} non-empty lines extracted")
        
        return text
        
    except Exception as e:
        print(f"❌ Error extracting text from image {image_path}: {e}")
        print(f"   Check if tesseract is properly configured: {pytesseract.pytesseract.tesseract_cmd}")
        return ""

def extract_rooms_from_gih_images(image_paths, schedule_date):
    """Extract and classify rooms from GIH image files using OCR
    Focus on extracting room numbers from the first column and their corresponding dates
    """
    try:
        all_lines = []
        
        # Extract text from all images and collect lines
        for image_path in image_paths:
            text = extract_text_from_image(image_path)
            lines = text.split('\n')
            all_lines.extend(lines)
        
        room_data = []
        
        for line in all_lines:
            line_clean = line.strip()
            if not line_clean or len(line_clean) < 10:  # Skip too short lines
                continue
            
            # First, try to find room number at the beginning of line (like PDF processing)
            room_match = re.match(r'^\s*(\d{4})\b', line_clean)
            if room_match:
                room_number = room_match.group(1)
                
                # Skip years (19xx, 20xx)
                if re.match(r'^(19|20)\d{2}$', room_number):
                    continue
                
                # Now extract dates from the same line
                # Look for patterns like "11-08-25" or "11/08/25" 
                dates_found = re.findall(r'\b(\d{2}[-/]\d{2}[-/]\d{2})\b', line_clean)
                
                # Convert different date formats to DD-MM-YY
                normalized_dates = []
                for date_str in dates_found:
                    # Handle both DD-MM-YY and DD/MM/YY formats
                    date_clean = date_str.replace('/', '-')
                    normalized_dates.append(date_clean)
                
                if len(normalized_dates) >= 2:
                    # Typically first date is check-in, second is check-out
                    checkin_date = normalized_dates[0]
                    checkout_date = normalized_dates[1]
                    
                    room_data.append({
                        'room': room_number,
                        'checkin': checkin_date,
                        'checkout': checkout_date,
                        'source_line': line_clean[:50]  # Keep first 50 chars for debugging
                    })
                elif len(normalized_dates) == 1:
                    # Single date - treat as current stay date
                    room_data.append({
                        'room': room_number,
                        'checkin': normalized_dates[0],
                        'checkout': normalized_dates[0],
                        'source_line': line_clean[:50]
                    })
                else:
                    # Room number found but no dates - might be header or incomplete line
                    print(f"Room {room_number} found but no dates in line: {line_clean[:30]}...")
            else:
                # Alternative approach: find any room numbers in the line with their nearby dates
                room_matches = re.findall(r'\b(\d{4})\b', line_clean)
                for room_number in room_matches:
                    # Skip years and common numbers
                    if re.match(r'^(19|20)\d{2}$', room_number):
                        continue
                    if room_number in ['1844', '1103']:  # Skip common time/reference numbers
                        continue
                        
                    # Extract dates from the same line 
                    dates_found = re.findall(r'\b(\d{2}[-/]\d{2}[-/]\d{2})\b', line_clean)
                    
                    if len(dates_found) >= 2:
                        normalized_dates = [date_str.replace('/', '-') for date_str in dates_found]
                        checkin_date = normalized_dates[0]
                        checkout_date = normalized_dates[1]
                        
                        room_data.append({
                            'room': room_number,
                            'checkin': checkin_date,
                            'checkout': checkout_date,
                            'source_line': line_clean[:50]
                        })
        
        # Remove duplicates based on room + dates combination
        seen_rooms = set()
        unique_room_data = []
        
        for data in room_data:
            room_key = f"{data['room']}_{data['checkin']}_{data['checkout']}"
            if room_key not in seen_rooms:
                seen_rooms.add(room_key)
                unique_room_data.append(data)
                print(f"Found: Room {data['room']}, CI: {data['checkin']}, CO: {data['checkout']}")
        
        # Classify rooms based on schedule date
        gih_arr_rooms = []
        gih_od_rooms = []
        
        for room_info in unique_room_data:
            room = room_info['room']
            checkin = room_info['checkin']
            checkout = room_info['checkout']
            
            if checkin == schedule_date:
                gih_arr_rooms.append(room)
                print(f"ARR: Room {room} (check-in on {checkin})")
            elif checkout == schedule_date:
                # Skip DEP rooms from GIH as mentioned in original logic
                print(f"Skip DEP: Room {room} (check-out on {checkout})")
                pass
            else:
                gih_od_rooms.append(room)
                print(f"OD: Room {room} (staying {checkin} to {checkout})")
        
        print(f"GIH Images processed: {len(gih_arr_rooms)} ARR, {len(gih_od_rooms)} OD")
        
        return {
            'ARR': sorted(list(set(gih_arr_rooms))),
            'OD': sorted(list(set(gih_od_rooms)))
        }
        
    except Exception as e:
        print(f"Error processing GIH images: {e}")
        import traceback
        traceback.print_exc()
        return {'ARR': [], 'OD': []}

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1] in ALLOWED_EXTENSIONS

def pdf_to_text(pdf_path):
    """Convert PDF thành text sử dụng pdfplumber (không cần pdftotext)"""
    text_path = pdf_path.replace('.pdf', '.txt').replace('.PDF', '.txt')
    
    try:
        # Try system pdftotext first
        cmd = ['pdftotext', '-layout', pdf_path, text_path]
        result = subprocess.run(cmd, capture_output=True, text=True, timeout=30)
        
        if result.returncode == 0 and os.path.exists(text_path):
            return text_path
    except:
        pass
    
    # Fallback to pdfplumber
    try:
        import pdfplumber
        
        with pdfplumber.open(pdf_path) as pdf:
            text_content = ""
            for page in pdf.pages:
                if page.extract_text():
                    text_content += page.extract_text() + "\n"
        
        # Write to text file
        with open(text_path, 'w', encoding='utf-8') as f:
            f.write(text_content)
        
        return text_path
        
    except Exception as e:
        print(f"Error extracting PDF text: {e}")
        return None

def extract_rooms_from_arr_dep(pdf_path):
    """Trích xuất số phòng từ file ARR/DEP - chỉ lấy từ cột đầu tiên"""
    text_path = pdf_to_text(pdf_path)
    if not text_path:
        return []
    
    try:
        with open(text_path, 'r', encoding='utf-8') as f:
            content = f.read()
        
        lines = content.split('\n')
        rooms = []
        
        for line in lines:
            line_clean = line.strip()
            if not line_clean:
                continue
                
            # Chỉ lấy số phòng ở đầu dòng (cột đầu tiên)
            # Pattern tìm số 4 chữ số ở đầu dòng, có thể có khoảng trắng phía trước
            room_match = re.match(r'^\s*(\d{4})\b', line_clean)
            
            if room_match:
                room = room_match.group(1)
                # Kiểm tra không phải năm (19xx hoặc 20xx)
                if not re.match(r'^(19|20)\d{2}$', room):
                    rooms.append(room)
        
        # Remove duplicates and sort
        unique_rooms = sorted(list(set(rooms)))
        
        # Clean up temp file
        os.remove(text_path)
        
        return unique_rooms
        
    except Exception as e:
        return []

def extract_rooms_from_gih(pdf_path, schedule_date):
    """Trích xuất và phân loại phòng từ file GIH"""
    text_path = pdf_to_text(pdf_path)
    if not text_path:
        return {'ARR': [], 'OD': []}
    
    try:
        with open(text_path, 'r', encoding='utf-8') as f:
            content = f.read()
        
        lines = content.split('\n')
        room_data = []
        
        for i, line in enumerate(lines):
            line_clean = line.strip()
            if not line_clean:
                continue
            
            room_match = re.match(r'^(\d{4})', line_clean)
            
            if room_match:
                room_number = room_match.group(1)
                dates_found = re.findall(r'\b(\d{2}-\d{2}-\d{2})\b', line_clean)
                
                if len(dates_found) >= 2:
                    checkin_date = dates_found[0]
                    checkout_date = dates_found[1]
                    
                    room_data.append({
                        'room': room_number,
                        'checkin': checkin_date,
                        'checkout': checkout_date
                    })
        
        # Remove duplicates
        seen_rooms = set()
        unique_room_data = []
        
        for data in room_data:
            room_key = f"{data['room']}_{data['checkin']}_{data['checkout']}"
            if room_key not in seen_rooms:
                seen_rooms.add(room_key)
                unique_room_data.append(data)
        
        # Classify rooms
        gih_arr_rooms = []
        gih_od_rooms = []
        
        for room_info in unique_room_data:
            room = room_info['room']
            checkin = room_info['checkin']
            checkout = room_info['checkout']
            
            if checkin == schedule_date:
                gih_arr_rooms.append(room)
            elif checkout == schedule_date:
                pass  # Skip DEP rooms from GIH
            else:
                gih_od_rooms.append(room)
        
        # Clean up temp file
        os.remove(text_path)
        
        return {
            'ARR': sorted(list(set(gih_arr_rooms))),
            'OD': sorted(list(set(gih_od_rooms)))
        }
        
    except Exception as e:
        return {'ARR': [], 'OD': []}

def create_excel_output(result, schedule_date):
    """Cập nhật template Excel với kết quả phân loại"""
    try:
        import openpyxl
        from datetime import datetime
        
        # Path to template and output
        template_path = 'template.xlsx'
        output_path = os.path.join(app.config['UPLOAD_FOLDER'], f'room_classification_{schedule_date.replace("-", "")}.xlsx')
        
        if not os.path.exists(template_path):
            print(f"Template file not found: {template_path}")
            return None
        
        # Load template
        wb = openpyxl.load_workbook(template_path)
        sheet = wb.active
        
        # Parse schedule date to update in template
        try:
            date_obj = datetime.strptime(schedule_date, "%d-%m-%y")
            formatted_date = date_obj.strftime("%d/%m/%Y")
        except:
            formatted_date = schedule_date
        
        # Update date in template (usually in row 3)
        for row in range(1, 10):
            for col in range(1, 10):
                cell_value = sheet.cell(row=row, column=col).value
                if cell_value and 'Date:' in str(cell_value):
                    sheet.cell(row=row, column=col, value=f'Date: {formatted_date}')
                    break
        
        # Find header sections (Room, OD, DO, ARR columns)
        header_sections = []  # List of {room_col, od_col, do_col, arr_col}
        header_row = 4  # Based on template analysis
        
        for col in range(1, sheet.max_column + 1):
            header_cell = sheet.cell(row=header_row, column=col)
            if header_cell.value:
                header_val = str(header_cell.value).strip()
                if header_val == 'Room':
                    # Found a room column, next columns should be OD, DO, ARR, NOTE
                    room_col = col
                    od_col = col + 1 if col + 1 <= sheet.max_column else None
                    do_col = col + 2 if col + 2 <= sheet.max_column else None
                    arr_col = col + 3 if col + 3 <= sheet.max_column else None
                    
                    # Verify column headers
                    od_header = sheet.cell(row=header_row, column=od_col).value if od_col else None
                    do_header = sheet.cell(row=header_row, column=do_col).value if do_col else None
                    arr_header = sheet.cell(row=header_row, column=arr_col).value if arr_col else None
                    
                    if (od_header and 'OD' in str(od_header) and 
                        do_header and 'DO' in str(do_header) and 
                        arr_header and 'ARR' in str(arr_header)):
                        
                        header_sections.append({
                            'room_col': room_col,
                            'od_col': od_col,
                            'do_col': do_col,
                            'arr_col': arr_col
                        })
        
        print(f"Found {len(header_sections)} header sections")
        
        # Convert result room numbers to integers for comparison
        arr_room_ints = set()
        for room in result['ARR']:
            try:
                arr_room_ints.add(int(room))
            except:
                pass
                
        dep_room_ints = set()
        for room in result['DEP']:
            try:
                dep_room_ints.add(int(room))
            except:
                pass
                
        od_room_ints = set()
        for room in result['OD']:
            try:
                od_room_ints.add(int(room))
            except:
                pass
        
        print(f"Room sets: ARR={len(arr_room_ints)}, DEP={len(dep_room_ints)}, OD={len(od_room_ints)}")
        
        # Mark rooms with X - process each row starting from row 5
        marked_rooms = {'ARR': 0, 'DEP': 0, 'OD': 0}
        
        for row_num in range(5, sheet.max_row + 1):
            for section in header_sections:
                room_cell = sheet.cell(row=row_num, column=section['room_col'])
                if room_cell.value:
                    try:
                        room_value = str(room_cell.value).strip()
                        if room_value.isdigit():
                            # Convert room number (handle formats like 0211 -> 211)
                            room_num = int(room_value.lstrip('0')) if room_value.startswith('0') else int(room_value)
                            
                            # Mark appropriate columns
                            if room_num in od_room_ints and section['od_col']:
                                sheet.cell(row=row_num, column=section['od_col'], value='X')
                                marked_rooms['OD'] += 1
                            
                            if room_num in dep_room_ints and section['do_col']:
                                sheet.cell(row=row_num, column=section['do_col'], value='X')
                                marked_rooms['DEP'] += 1
                            
                            if room_num in arr_room_ints and section['arr_col']:
                                sheet.cell(row=row_num, column=section['arr_col'], value='X')
                                marked_rooms['ARR'] += 1
                                
                    except ValueError:
                        continue  # Skip non-numeric room values
        
        print(f"Marked rooms: {marked_rooms}")
        
        # Add summary totals to EA:, DO:, OD: cells (row 38)
        # EA: = ARR total (column G, next to F38)
        # DO: = DEP total (column I, next to H38) 
        # OD: = OD total (column K, next to J38)
        try:
            # Check if manual totals are provided in result
            ea_total = result.get('manual_ea', len(arr_room_ints))
            do_total = result.get('manual_do', len(dep_room_ints))
            od_total = result.get('manual_od', len(od_room_ints))
            
            sheet.cell(row=38, column=7, value=ea_total)  # G38: EA total
            sheet.cell(row=38, column=9, value=do_total)  # I38: DO total
            sheet.cell(row=38, column=11, value=od_total)  # K38: OD total
            print(f"Added totals: EA={ea_total}, DO={do_total}, OD={od_total}")
        except Exception as e:
            print(f"Error adding totals: {e}")
        
        # Save the updated Excel
        wb.save(output_path)
        
        return output_path
        
    except Exception as e:
        print(f"Error creating Excel: {e}")
        import traceback
        traceback.print_exc()
        return None

def get_compdf_access_token():
    """Get access token from ComPDF API"""
    import requests
    
    PUBLIC_KEY = os.environ.get('COMPDF_PUBLIC_KEY')
    SECRET_KEY = os.environ.get('COMPDF_SECRET_KEY')
    
    if not PUBLIC_KEY or not SECRET_KEY:
        print("ERROR: ComPDF API credentials not found in environment variables")
        return None
    
    try:
        token_url = "https://api-server.compdf.com/server/v1/oauth/token"
        
        data = {
            "publicKey": PUBLIC_KEY,
            "secretKey": SECRET_KEY
        }
        
        headers = {
            "Content-Type": "application/json"
        }
        
        response = requests.post(token_url, json=data, headers=headers, timeout=30)
        
        if response.status_code == 200:
            result = response.json()
            if 'data' in result and 'accessToken' in result['data']:
                return result['data']['accessToken']
        
        print(f"Failed to get access token: {response.text}")
        return None
        
    except Exception as e:
        print(f"Error getting access token: {e}")
        return None

def convert_excel_to_pdf_via_compdf(excel_path):
    """Convert Excel to PDF using ComPDF API with correct workflow from documentation"""
    import requests
    import json
    import time
    
    # Get access token
    access_token = get_compdf_access_token()
    if not access_token:
        print("Could not get ComPDF access token")
        return None
    
    try:
        headers = {
            "Authorization": f"Bearer {access_token}"
        }
        
        # Step 1: Get tool support list to find the correct executeTypeUrl
        print("Getting ComPDF tool support list...")
        tools_url = "https://api-server.compdf.com/server/v1/tool/support"
        tools_response = requests.get(tools_url, headers=headers, timeout=30)
        
        if tools_response.status_code != 200:
            print(f"Failed to get tool support: {tools_response.text}")
            return None
            
        tools_result = tools_response.json()
        if tools_result.get('code') != '200':
            print(f"Tool support request failed: {tools_result}")
            return None
            
        # Find Excel to PDF conversion tool
        execute_type_url = None
        for tool in tools_result.get('data', []):
            if (tool.get('sourceTypeName') == 'xlsx' and 
                tool.get('targetTypeName') == 'pdf'):
                execute_type_url = tool.get('executeTypeUrl')
                break
        
        # Fallback to xls format if xlsx not found
        if not execute_type_url:
            for tool in tools_result.get('data', []):
                if (tool.get('sourceTypeName') == 'xls' and 
                    tool.get('targetTypeName') == 'pdf'):
                    execute_type_url = tool.get('executeTypeUrl')
                    break
        
        if not execute_type_url:
            # Fallback to common patterns if exact match not found
            for tool in tools_result.get('data', []):
                if 'excel' in tool.get('executeTypeUrl', '').lower() or 'office' in tool.get('executeTypeUrl', '').lower():
                    if 'pdf' in tool.get('executeTypeUrl', '').lower():
                        execute_type_url = tool.get('executeTypeUrl')
                        break
        
        if not execute_type_url:
            print("Could not find Excel to PDF conversion tool in supported tools")
            print(f"Available tools: {[tool.get('executeTypeUrl') for tool in tools_result.get('data', [])[:10]]}")
            return None
            
        print(f"Using executeTypeUrl: {execute_type_url}")
        
        # Step 2: Create task using GET with executeTypeUrl as path parameter
        print("Creating ComPDF conversion task...")
        create_task_url = f"https://api-server.compdf.com/server/v1/task/{execute_type_url}?language=1"
        create_response = requests.get(create_task_url, headers=headers, timeout=30)
        
        if create_response.status_code != 200:
            print(f"Failed to create task: {create_response.text}")
            return None
            
        create_result = create_response.json()
        if create_result.get('code') != '200':
            print(f"Task creation failed: {create_result}")
            return None
            
        task_id = create_result['data']['taskId']
        print(f"Task created: {task_id}")
        
        # Step 2: Upload file
        upload_url = "https://api-server.compdf.com/server/v1/file/upload"
        upload_headers = {
            "Authorization": f"Bearer {access_token}"
        }
        
        with open(excel_path, 'rb') as f:
            files = {
                'file': (os.path.basename(excel_path), f, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            }
            upload_data = {
                'taskId': task_id,
                'language': '1'  # 1 for English, 2 for Chinese
            }
            
            print("Uploading Excel file...")
            upload_response = requests.post(upload_url, headers=upload_headers, files=files, data=upload_data, timeout=60)
        
        if upload_response.status_code != 200:
            print(f"Failed to upload file: {upload_response.text}")
            return None
            
        upload_result = upload_response.json()
        if upload_result.get('code') != '200':
            print(f"File upload failed: {upload_result}")
            return None
            
        file_key = upload_result['data']['fileKey']
        print(f"File uploaded: {file_key}")
        
        # Step 3: Execute conversion using GET with query parameters
        print("Starting conversion...")
        execute_url = f"https://api-server.compdf.com/server/v1/execute/start?taskId={task_id}&language=1"
        execute_headers = {
            "Authorization": f"Bearer {access_token}"
        }
        
        execute_response = requests.get(execute_url, headers=execute_headers, timeout=30)
        
        if execute_response.status_code != 200:
            print(f"Failed to execute conversion: {execute_response.text}")
            return None
            
        execute_result = execute_response.json()
        if execute_result.get('code') != '200':
            print(f"Conversion execution failed: {execute_result}")
            return None
            
        print("Conversion started, waiting for completion...")
        
        # Step 4: Check status and download
        max_attempts = 30  # Wait up to 5 minutes
        for attempt in range(max_attempts):
            time.sleep(10)  # Wait 10 seconds between checks
            
            print(f"Checking status... (attempt {attempt + 1}/{max_attempts})")
            
            # Check file info to get status
            status_url = f"https://api-server.compdf.com/server/v1/file/fileInfo?fileKey={file_key}&language=1"
            status_headers = {
                "Authorization": f"Bearer {access_token}"
            }
            
            status_response = requests.get(status_url, headers=status_headers, timeout=30)
            
            if status_response.status_code != 200:
                print(f"Failed to check status: {status_response.text}")
                continue
                
            status_result = status_response.json()
            if status_result.get('code') != '200':
                print(f"Status check failed: {status_result}")
                continue
                
            status_data = status_result.get('data', {})
            task_status = status_data.get('status')
            print(f"Task status: {task_status}")
            
            if task_status in ['TaskFinish', 'success']:
                # Download the converted PDF
                download_url = status_data.get('downloadUrl')
                if not download_url:
                    print("No download URL found")
                    return None
                    
                print(f"Downloading PDF from: {download_url}")
                
                download_response = requests.get(download_url, timeout=60)
                if download_response.status_code == 200:
                    return download_response.content
                else:
                    print(f"Failed to download PDF: {download_response.text}")
                    return None
                    
            elif task_status in ['TaskFail', 'TaskError']:
                failure_reason = status_data.get('failureReason', 'Unknown error')
                failure_code = status_data.get('failureCode', 'Unknown code')
                print(f"Task failed: Code={failure_code}, Reason={failure_reason}")
                return None
            elif task_status == 'TaskProcessing':
                continue  # Keep waiting
                
        print("Timeout waiting for conversion to complete")
        return None
        
    except Exception as e:
        print(f"ComPDF API error: {e}")
        import traceback
        traceback.print_exc()
        return None

def create_image_from_excel(excel_path):
    """Tạo ảnh từ file Excel sử dụng ComPDF API"""
    try:
        image_path = excel_path.replace('.xlsx', '.png')
        
        # Step 1: Convert Excel to PDF using ComPDF API
        print("Converting Excel to PDF using ComPDF API...")
        pdf_data = convert_excel_to_pdf_via_compdf(excel_path)
        
        if not pdf_data:
            print("Failed to convert Excel to PDF via ComPDF API")
            return None
            
        # Step 2: Save PDF temporarily and convert to PNG
        with tempfile.TemporaryDirectory() as temp_dir:
            temp_pdf_path = os.path.join(temp_dir, 'converted.pdf')
            
            # Write PDF data to temporary file
            with open(temp_pdf_path, 'wb') as f:
                f.write(pdf_data)
                
            print(f"PDF saved to {temp_pdf_path}, converting to image...")
            
            # Step 3: Convert PDF to PNG using pdf2image
            images = convert_from_path(temp_pdf_path, dpi=200, fmt='PNG', first_page=1, last_page=1)
            
            if not images:
                print("Failed to convert PDF to image")
                return None
                
            image = images[0]
            if image.mode != 'RGB':
                image = image.convert('RGB')
                
            # Crop whitespace
            bbox = image.getbbox()
            if bbox:
                padding = 20
                left, top, right, bottom = bbox
                left = max(0, left - padding)
                top = max(0, top - padding)  
                right = min(image.width, right + padding)
                bottom = min(image.height, bottom + padding)
                
                image = image.crop((left, top, right, bottom))
            
            image.save(image_path, 'PNG', quality=95, optimize=True)
            print(f"Image saved to {image_path}")
            
            return image_path
            
    except Exception as e:
        print(f"Error creating image: {e}")
        import traceback
        traceback.print_exc()
        return None

@app.route('/download/<path:filename>')
def download_file(filename):
    """Download generated files"""
    try:
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        if os.path.exists(file_path):
            return send_file(file_path, as_attachment=True)
        else:
            return jsonify({'error': 'File not found'}), 404
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/preview/<path:filename>')
def preview_file(filename):
    """Preview generated image files inline"""
    try:
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        if os.path.exists(file_path) and filename.lower().endswith(('.png', '.jpg', '.jpeg', '.gif')):
            return send_file(file_path, as_attachment=False)
        else:
            return jsonify({'error': 'Image file not found'}), 404
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/')
def index():
    return render_template('upload.html')

@app.route('/debug', methods=['GET', 'POST'])
def debug_endpoint():
    """Debug endpoint to test basic functionality"""
    try:
        if request.method == 'GET':
            return jsonify({
                'status': 'ok',
                'method': 'GET',
                'timestamp': datetime.now().isoformat(),
                'environment': {
                    'has_compdf_keys': bool(os.environ.get('COMPDF_PUBLIC_KEY')),
                    'upload_folder': app.config['UPLOAD_FOLDER'],
                    'upload_folder_exists': os.path.exists(app.config['UPLOAD_FOLDER']),
                    'tesseract_cmd': pytesseract.pytesseract.tesseract_cmd
                }
            })
        
        elif request.method == 'POST':
            return jsonify({
                'status': 'ok',
                'method': 'POST',
                'timestamp': datetime.now().isoformat(),
                'form_keys': list(request.form.keys()),
                'files_keys': list(request.files.keys()),
                'content_type': request.content_type
            })
            
    except Exception as e:
        return jsonify({
            'status': 'error',
            'error': str(e),
            'timestamp': datetime.now().isoformat()
        }), 500

@app.route('/manual_edit', methods=['GET', 'POST'])
def manual_edit():
    if request.method == 'GET':
        # Get data from session or query parameters
        data = request.args.to_dict()
        if 'ARR' in data:
            data['ARR'] = data['ARR'].split(',') if data['ARR'] else []
        if 'DEP' in data:
            data['DEP'] = data['DEP'].split(',') if data['DEP'] else []
        if 'OD' in data:
            data['OD'] = data['OD'].split(',') if data['OD'] else []
        
        return render_template('manual_edit.html', data=data)
    
    elif request.method == 'POST':
        # Process manual edits and create final Excel
        try:
            schedule_date = request.form.get('schedule_date', '')
            
            # Get manually edited room lists
            arr_rooms_str = request.form.get('arr_rooms', '').strip()
            dep_rooms_str = request.form.get('dep_rooms', '').strip() 
            od_rooms_str = request.form.get('od_rooms', '').strip()
            
            # Parse room lists
            arr_rooms = [room.strip() for room in arr_rooms_str.split(',') if room.strip()] if arr_rooms_str else []
            dep_rooms = [room.strip() for room in dep_rooms_str.split(',') if room.strip()] if dep_rooms_str else []
            od_rooms = [room.strip() for room in od_rooms_str.split(',') if room.strip()] if od_rooms_str else []
            
            # Get manual totals (if provided)
            manual_ea = request.form.get('manual_ea', '').strip()
            manual_do = request.form.get('manual_do', '').strip()
            manual_od = request.form.get('manual_od', '').strip()
            
            result = {
                'schedule_date': schedule_date,
                'ARR': arr_rooms,
                'DEP': dep_rooms,
                'OD': od_rooms,
                'processing_info': [f'Manual edit: ARR={len(arr_rooms)}, DEP={len(dep_rooms)}, OD={len(od_rooms)}']
            }
            
            # Add manual totals if provided
            if manual_ea and manual_ea.isdigit():
                result['manual_ea'] = int(manual_ea)
                result['processing_info'].append(f'Manual EA total: {manual_ea}')
            
            if manual_do and manual_do.isdigit():
                result['manual_do'] = int(manual_do)
                result['processing_info'].append(f'Manual DO total: {manual_do}')
            
            if manual_od and manual_od.isdigit():
                result['manual_od'] = int(manual_od)
                result['processing_info'].append(f'Manual OD total: {manual_od}')
            
            # Create Excel file with manually edited results
            excel_path = create_excel_output(result, schedule_date)
            if excel_path:
                result['excel_path'] = excel_path
                
                # Create image from Excel
                image_path = create_image_from_excel(excel_path)
                if image_path:
                    result['image_path'] = image_path
            
            return jsonify(result)
            
        except Exception as e:
            return jsonify({'error': f'Lỗi xử lý: {str(e)}'}), 500

@app.after_request
def after_request(response):
    """Add CORS headers to all responses"""
    response.headers.add('Access-Control-Allow-Origin', '*')
    response.headers.add('Access-Control-Allow-Headers', 'Content-Type,Authorization')
    response.headers.add('Access-Control-Allow-Methods', 'GET,PUT,POST,DELETE,OPTIONS')
    return response

@app.route('/upload', methods=['POST', 'OPTIONS'])
def upload_files():
    # Handle preflight OPTIONS request
    if request.method == 'OPTIONS':
        return '', 200
        
    print("\n🚀 === UPLOAD REQUEST RECEIVED ===")
    print(f"Request method: {request.method}")
    print(f"Content-Type: {request.content_type}")
    print(f"Form data keys: {list(request.form.keys())}")
    print(f"Files: {list(request.files.keys())}")
    
    try:
        schedule_date = request.form.get('schedule_date', '')
        print(f"📅 Schedule date received: '{schedule_date}'")
        
        if not schedule_date:
            print("❌ No schedule date provided")
            return jsonify({'error': 'Vui lòng nhập ngày chia lịch'}), 400
        
        # Validate date format
        try:
            datetime.strptime(schedule_date, "%d-%m-%y")
        except ValueError:
            return jsonify({'error': 'Format ngày không đúng (DD-MM-YY)'}), 400
            
        # Check if at least one file is provided
        has_files = False
        for file_key in ['arr_file', 'dep_file', 'gih_file']:
            if file_key in request.files and request.files[file_key] and request.files[file_key].filename:
                has_files = True
                break
                
        if not has_files:
            return jsonify({'error': 'Vui lòng upload ít nhất 1 file'}), 400
        
        result = {
            'schedule_date': schedule_date,
            'ARR': [],
            'DEP': [],
            'OD': [],
            'processing_info': []
        }
        
        # Process ARR file
        if 'arr_file' in request.files:
            arr_file = request.files['arr_file']
            if arr_file and arr_file.filename and allowed_file(arr_file.filename):
                filename = secure_filename(arr_file.filename)
                filepath = os.path.join(app.config['UPLOAD_FOLDER'], 'arr_' + filename)
                arr_file.save(filepath)
                
                arr_rooms = extract_rooms_from_arr_dep(filepath)
                result['ARR'] = arr_rooms
                result['processing_info'].append(f"ARR: {len(arr_rooms)} phòng từ {filename}")
                
                os.remove(filepath)  # Clean up
        
        # Process DEP file
        if 'dep_file' in request.files:
            dep_file = request.files['dep_file']
            if dep_file and dep_file.filename and allowed_file(dep_file.filename):
                filename = secure_filename(dep_file.filename)
                filepath = os.path.join(app.config['UPLOAD_FOLDER'], 'dep_' + filename)
                dep_file.save(filepath)
                
                dep_rooms = extract_rooms_from_arr_dep(filepath)
                result['DEP'] = dep_rooms
                result['processing_info'].append(f"DEP: {len(dep_rooms)} phòng từ {filename}")
                
                os.remove(filepath)  # Clean up
        
        # Process GIH files (PDF or multiple images)
        if 'gih_file' in request.files:
            gih_files = request.files.getlist('gih_file')  # Get all files with this name
            
            if gih_files and gih_files[0].filename:  # At least one file exists
                gih_result = {'ARR': [], 'OD': []}
                
                # Check if it's a single PDF file
                if len(gih_files) == 1 and allowed_file(gih_files[0].filename):
                    # Single PDF file
                    gih_file = gih_files[0]
                    filename = secure_filename(gih_file.filename)
                    filepath = os.path.join(app.config['UPLOAD_FOLDER'], 'gih_' + filename)
                    gih_file.save(filepath)
                    
                    gih_result = extract_rooms_from_gih(filepath, schedule_date)
                    result['processing_info'].append(f"GIH PDF: {len(gih_result['OD'])} OD phòng, {len(gih_result['ARR'])} thêm vào ARR từ {filename}")
                    
                    os.remove(filepath)  # Clean up
                    
                else:
                    # Multiple image files or single image file
                    image_paths = []
                    saved_files = []
                    
                    try:
                        for i, gih_file in enumerate(gih_files):
                            if gih_file.filename and allowed_image_file(gih_file.filename):
                                filename = secure_filename(gih_file.filename)
                                filepath = os.path.join(app.config['UPLOAD_FOLDER'], f'gih_{i}_{filename}')
                                gih_file.save(filepath)
                                image_paths.append(filepath)
                                saved_files.append(filepath)
                        
                        if image_paths:
                            print(f"Processing {len(image_paths)} GIH image files...")
                            gih_result = extract_rooms_from_gih_images(image_paths, schedule_date)
                            result['processing_info'].append(f"GIH Images: {len(image_paths)} ảnh, {len(gih_result['OD'])} OD phòng, {len(gih_result['ARR'])} thêm vào ARR")
                        else:
                            result['processing_info'].append(f"Không tìm thấy ảnh GIH hợp lệ trong {len(gih_files)} files")
                            
                    except Exception as e:
                        print(f"Error processing GIH image files: {e}")
                        result['processing_info'].append(f"Lỗi xử lý ảnh GIH: {str(e)}")
                        
                    finally:
                        # Clean up saved image files
                        for filepath in saved_files:
                            try:
                                os.remove(filepath)
                            except:
                                pass
                
                # Merge ARR from GIH with ARR from file
                if gih_result['ARR'] or gih_result['OD']:
                    combined_arr = list(set(result['ARR'] + gih_result['ARR']))
                    combined_arr.sort()
                    result['ARR'] = combined_arr
                    
                    result['OD'] = gih_result['OD']
        
        # Create Excel file with results
        excel_path = create_excel_output(result, schedule_date)
        if excel_path:
            result['excel_path'] = excel_path
            
            # Create image from Excel (optional - graceful fallback if LibreOffice unavailable)
            try:
                image_path = create_image_from_excel(excel_path)
                if image_path:
                    result['image_path'] = image_path
                    result['processing_info'].append("✅ Excel và ảnh đã được tạo thành công")
                else:
                    result['processing_info'].append("⚠️ Excel đã tạo thành công, nhưng không thể tạo ảnh (LibreOffice không khả dụng)")
            except Exception as e:
                result['processing_info'].append(f"⚠️ Excel đã tạo thành công, lỗi tạo ảnh: {str(e)}")
                print(f"Image creation error: {e}")
        
        print(f"\n✅ === PROCESSING COMPLETE ===")
        print(f"Final result keys: {list(result.keys())}")
        print(f"Result summary: ARR={len(result['ARR'])}, DEP={len(result['DEP'])}, OD={len(result['OD'])}")
        print(f"Excel path: {result.get('excel_path', 'None')}")
        print(f"Image path: {result.get('image_path', 'None')}")
        print(f"Processing info: {result['processing_info']}")
        
        return jsonify(result)
        
    except Exception as e:
        print(f"\n❌ === UPLOAD ERROR ===")
        print(f"Error: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': f'Lỗi xử lý: {str(e)}'}), 500

if __name__ == '__main__':
    # Get port from environment variable (for Render deployment)
    port = int(os.environ.get('PORT', 8000))
    
    print("🏨 Starting Hotel Room Classification Web Server...")
    print(f"📡 Server will be available on port {port}")
    print("📋 PDF processing: pdftotext + ComPDF API for image generation")
    
    # Use different settings for production vs development
    debug_mode = os.environ.get('FLASK_ENV') != 'production'
    app.run(debug=debug_mode, host='0.0.0.0', port=port)
